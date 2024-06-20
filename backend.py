import pandas as pd
from pdfminer.high_level import extract_text
import pdfplumber
import re
import openpyxl
from openpyxl import load_workbook
from mutliclass import convert_to_df_multiclass, multiclass_count
import os


# =====================
# extract text from pdf
# =====================
"""
input: pdf
output: text extracted from pdf
"""


def extract_text_from_pdf(pdf):
    pdf_folder = "submission_pdfs/"
    txt_folder = "text_files/"
    pdf_text = ""

    # Extract the text
    with pdfplumber.open(pdf) as pdf:
        for page in pdf.pages:
            pdf_text += page.extract_text()
    return pdf_text


# =============
# extract table
# =============
"""
takes text and outputs the tables into dataframes 
    # of columns in df is dependent on the information in the submission
    ex. if submission contains "paid" column so will the df
        if submission does not contain "paid" the df will not have a "paid" column
        the number of columns will vary based on submission
input: text
output: df containing tables
"""


def extract_table_from_text(pdf_text):
    lower_text = pdf_text.lower()
    text_list = lower_text.split("\n")

    start_index = -1
    val_index = -1

    val_date = ""

    year_pattern = "\d{2}/\d{2}/\d{4}-\d{2}/\d{2}/\d{4}"
    eval_year_pattern = "\d{2}/\d{2}/\d{4}"
    for index, item in enumerate(text_list, start=1):
        # get index
        if re.match("Valuation Date:", item):
            val_index = index
        if re.match(year_pattern, item):
            start_index = index
            break
    
    # Gets the evaluation year
    for item in text_list[val_index - 1].split(" "):
        if re.match(eval_year_pattern, item):
            val_date = item

    data_temp = text_list[start_index - 1 :]
    data = []
    for line in data_temp:
        if not re.match(year_pattern, line):
            break
        else:
            data.append(line)

    data.insert(0, text_list[start_index - 2])

    # cleaing columns
    columns = []
    for line in data:
        if re.match("year", line):
            # creating columns from line
            columns = line.split(" ")
            # fixing claims column
            for i in range(len(columns) - 1):
                if columns[i] == "#":
                    new_col = columns[i] + " " + columns[i + 1]
                    columns.pop(i)
                    columns.pop(i)
                    columns.insert(i, new_col)
                    break

    columns.pop(0)
    columns.insert(0, "end_date")
    columns.insert(0, "start_date")

    row_data = []

    for line in data:
        row = []
        if re.match(year_pattern, line):
            row = line.split(" ")
            for i in range(len(row)):
                if re.match(year_pattern, row[i]):
                    years = row[i].split("-")
                    row.pop(i)
                    row.insert(0, years[1])
                    row.insert(0, years[0])
                    row.insert(2, val_date)
                    break
            row_data.append(row)

    columns.insert(2, "eval_date")
        
    df = pd.DataFrame(row_data, columns=columns)
    print(df)
    return df


# ==================
# import df to excel
# ==================

"""
copies the Loss Rater template into a new file saved as the given filename
imports data from df into new Loss Rater and saves the file
takes a dataframe
input: df, name of file 
output: None
"""


def import_to_excel(df, file_name):
    # Load the original Loss Rater template
    original_file = "Loss Experience Template.xlsx"
    print("Loading origianl workbook...")
    origianl_workbook = load_workbook(filename=original_file)
    print("Workbook successfully loaded.\n")

    # Save the workbook as a new rater and close orginal
    print("Copying file...")
    origianl_workbook.save(filename=file_name + ".xlsx")
    print(f"File copied as {file_name}.xlsx.\n")
    origianl_workbook.close()

    # Loading Copy of the Loss Rater
    print("Loading copied filed...")
    wb = load_workbook(filename=file_name + ".xlsx")
    print("Copied file loaded.\n")
    ws = wb["Sheet1"]

    # target cells
    START_DATE_TARGET = "C9"
    END_DATE_TARGET = "D9"
    EVAL_DATE_TARGET = "E9"
    INCURRED_LOSSES_TARGET = "G9"
    PAID_LOSSES_TARGET = "I9"
    CLAIMS_COUNT_TARGET = "K9"

    # Starting at each target, load data with corresponding column
    #   after each insert, set "target" to the next cell in the column
    for i in range(df.shape[0]):
        ws[START_DATE_TARGET].value = df["start_date"][i]
        START_DATE_TARGET = ws.cell(
            row=ws[START_DATE_TARGET].row + 1, column=ws[START_DATE_TARGET].column
        ).coordinate

        ws[END_DATE_TARGET].value = df["end_date"][i]
        END_DATE_TARGET = ws.cell(
            row=ws[END_DATE_TARGET].row + 1, column=ws[END_DATE_TARGET].column
        ).coordinate

        ws[EVAL_DATE_TARGET].value = df["eval_date"][i]
        EVAL_DATE_TARGET = ws.cell(
            row=ws[EVAL_DATE_TARGET].row + 1, column=ws[EVAL_DATE_TARGET].column
        ).coordinate

        ws[CLAIMS_COUNT_TARGET].value = (
            df["# claims"][i] if "# claims" in df.columns else df["number"][i]
        )
        CLAIMS_COUNT_TARGET = ws.cell(
            row=ws[CLAIMS_COUNT_TARGET].row + 1, column=ws[CLAIMS_COUNT_TARGET].column
        ).coordinate

        ws[INCURRED_LOSSES_TARGET].value = df["incurred"][i]
        INCURRED_LOSSES_TARGET = ws.cell(
            row=ws[INCURRED_LOSSES_TARGET].row + 1,
            column=ws[INCURRED_LOSSES_TARGET].column,
        ).coordinate

        if "paid" in df.columns:
            ws[PAID_LOSSES_TARGET].value = df["paid"][i]
        elif "indemnity" in df.columns:
            ws[PAID_LOSSES_TARGET].value = df["indemnity"][i]

        PAID_LOSSES_TARGET = ws.cell(
            row=ws[PAID_LOSSES_TARGET].row + 1, column=ws[PAID_LOSSES_TARGET].column
        ).coordinate

    # Save completed Excel Rater
    print("Saving new file...")
    wb.save(filename=file_name + ".xlsx")
    print(file_name+".xlsx saved!")


# pdf = (
#     "submission_pdfs/Olympic Steel_ Inc__Submission_AL_2024-04-03_211242_95.pdf"
# )
# pdf_text = extract_text_from_pdf(pdf)
# file_name = pdf.split("/")[-1].split(" ")[0].strip("_") + " - Loss"
# num_classes, _ = multiclass_count(pdf_text)
# IS_MULTICLASS = num_classes > 1

# if IS_MULTICLASS:
#     df_list, fn_list = convert_to_df_multiclass(pdf, file_name)
#     for df, fn in zip(df_list, fn_list):
#         import_to_excel(df, fn)
# else:
#     df = extract_table_from_text(pdf_text)
#     import_to_excel(df, file_name)
pdf = "submission_pdfs/Fernlea Industries_ Inc__Submission_UMB_2024-06-04_012751_392.pdf"

def add_eval_date(eval_date, df_list):
    for df in df_list:
        eval_list = [eval_date] * df.shape[0]
        df["eval_date"] = eval_list


def start_backend(pdf):
    '''
    input: pdf(str) path to the submission pdf
    '''
    pdf_text = extract_text_from_pdf(pdf)
    file_name = pdf.split("/")[-1].split(" ")[0].strip("_")
    num_classes, _ = multiclass_count(pdf_text)
    IS_MULTICLASS = num_classes > 1

    if IS_MULTICLASS:
        df_list, fn_list, eval_date = convert_to_df_multiclass(pdf, file_name)
        add_eval_date(eval_date, df_list)
        for df, fn in zip(df_list, fn_list):
            print(df)
            import_to_excel(df, fn)

    else:
        df = extract_table_from_text(pdf_text)
        import_to_excel(df, file_name)
        os.system(f"start EXCEL.EXE {file_name}.xlsx")

    

start_backend(pdf)